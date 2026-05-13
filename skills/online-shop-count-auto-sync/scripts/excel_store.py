from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import WORKBOOK_PATH

SHEET_NAME = "在线店铺数"
HEADERS = ["日期", "美团在线店铺数", "饿了么在线店铺数", "采集时间"]


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(HEADERS)
    workbook.save(path)


def append_online_shop_row(
    stat_date: str,
    meituan_count: int,
    eleme_count: int,
    captured_at_text: str,
) -> Path:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not WORKBOOK_PATH.exists():
        _create_workbook(WORKBOOK_PATH)

    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    if worksheet.max_row == 1 and worksheet["A1"].value != HEADERS[0]:
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(HEADERS)

    worksheet.append([stat_date, meituan_count, eleme_count, captured_at_text])
    workbook.save(WORKBOOK_PATH)
    return WORKBOOK_PATH
